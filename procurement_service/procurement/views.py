from django.shortcuts import render, redirect, get_object_or_404
from .models import Product, Supplier, Price, SupplierToken, Category
from .forms import *
from datetime import datetime, time  # Правильный импорт
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.utils.dateparse import parse_date
import openpyxl
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone  # Для работы с часовыми поясами
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
import re
from django.urls import reverse

# Функция для создания границ
def get_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Функция для создания и применения стилей
def apply_styles(sheet):
    best_price_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = get_border()

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if cell.value == 'Best Price':
                cell.fill = best_price_fill

# Экспорт в Excel
@login_required
def export_to_excel(request):
    products = Product.objects.all()
    suppliers = Supplier.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Таблица продуктов"
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1

    headers = ["Наименование", "Средняя потребность", "Единица измерения"]
    for supplier in suppliers:
        headers.extend([f"{supplier.name} (Цена)", f"{supplier.name} (Производитель)"])
    sheet.append(headers)

    for product in products:
        row_data = [product.name, product.quantity, product.unit]
        price_cells = {}
        best_price = None

        for supplier in suppliers:
            price_entry = Price.objects.filter(product=product, supplier=supplier).first()
            price = price_entry.price if price_entry else None
            manufacturer = price_entry.manufacturer if price_entry else "-"

            row_data.append(price if price else "-")
            row_data.append(manufacturer)

            if price and (best_price is None or price < best_price):
                best_price = price
                price_cells[supplier.name] = price

        sheet.append(row_data)

        row_index = sheet.max_row
        col_index = 4
        for supplier in suppliers:
            if price_cells.get(supplier.name) == best_price:
                sheet.cell(row=row_index, column=col_index).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            col_index += 2

    apply_styles(sheet)

    for col_num, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2

    current_date = datetime.now().strftime("%Y-%m-%d")
    filename = f"products_{current_date}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response

@login_required
def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Чтение Excel-файла
                df = pd.read_excel(file)

                # Получаем список всех названий товаров из файла
                uploaded_product_names = df['Наименование товара'].tolist()

                # Удаляем продукты, которых нет в файле
                Product.objects.exclude(name__in=uploaded_product_names).delete()

                # Обновляем или создаем продукты
                for index, row in df.iterrows():
                    Product.objects.update_or_create(
                        name=row['Наименование товара'],  # Фильтр по названию
                        defaults={  # Обновляемые данные
                            'quantity': row['Средняя потребность на месяц'],
                            'unit': row['Единица измерения']
                        }
                    )

                return redirect('product_supplier_table')  # Перенаправление на таблицу
            except Exception as e:
                return render(request, 'excel/upload_excel.html', {
                    'form': form,
                    'error': f"Ошибка при обработке файла: {str(e)}"
                })
    else:
        form = UploadFileForm()
    return render(request, 'excel/upload_excel.html', {'form': form})

def user_login(request):
    # ✅ Если пользователь уже авторизован — сразу редирект
    if request.user.is_authenticated:
        return redirect('product_supplier_table')

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('product_supplier_table')
        else:
            return render(request, 'login/login.html', {'error': 'Неверное имя пользователя или пароль'})
    return render(request, 'login/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')  # Перенаправление на главную страницу

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Автоматически входить после регистрации
            return redirect('home')  # Перенаправление на главную страницу
    else:
        form = UserCreationForm()
    return render(request, 'login/register.html', {'form': form})

@login_required
def home(request):
    return render(request, 'home.html')

@login_required
def product_list(request):
    products = Product.objects.all()  # Получаем все продукты из базы данных
    return render(request, 'product/product_list.html', {'products': products})

@login_required
def add_product(request):
    if request.method == 'POST':
        # Если форма была отправлена, обрабатываем данные
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()  # Сохраняем продукт в базу данных
            return redirect('product_list')  # Перенаправляем на страницу со списком продуктов
    else:
        # Если это GET-запрос, показываем пустую форму
        form = ProductForm()
    
    # Рендерим шаблон с формой
    return render(request, 'product/product_card.html', {'form': form})

@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)  # Получаем продукт по ID
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)  # Заполняем форму данными продукта
        if form.is_valid():
            form.save()  # Сохраняем изменения
            return redirect('product_list')  # Перенаправляем на список продуктов
    else:
        form = ProductForm(instance=product)  # Показываем форму с текущими данными продукта
    return render(request, 'product/product_card.html', {'form': form})

@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)  # Получаем продукт по ID
    if request.method == 'POST':
        product.delete()  # Удаляем продукт
        return redirect('product_list')  # Перенаправляем на список продуктов
    return render(request, 'product/confirm_delete_product.html', {'product': product})

@login_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'category/category_list.html', {'categories': categories})


@login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')  # Убедись, что такой маршрут есть
    else:
        form = CategoryForm()
    return render(request, 'category/category_card.html', {'form': form})


@login_required
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'category/category_card.html', {'form': form})


@login_required
def delete_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'category/confirm_delete_category.html', {'category': category})

def toggle_supplier_visibility(request, supplier_id):
    if request.method == 'POST':
        try:
            supplier = Supplier.objects.get(id=supplier_id)
            data = json.loads(request.body)
            supplier.is_hidden = data.get('is_hidden', False)
            supplier.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required
def supplier_list(request):
    category_id = request.GET.get('category')  # получаем ID категории из запроса
    categories = Category.objects.all()

    show_hidden = request.GET.get('show_hidden', '0') == '1'

    query = request.GET.get('q')
    
    suppliers = Supplier.objects.all().order_by('name')
    if category_id:
        suppliers = suppliers.filter(category_id=category_id)

    if query:
        suppliers = suppliers.filter(
            Q(name__icontains=query) | Q(inn__icontains=query)
        )

    return render(request, 'supplier/supplier_list.html', {
        'suppliers': suppliers,
        'categories': categories,
        'selected_category': int(category_id) if category_id else None,
        'show_hidden': show_hidden,
    })

@login_required
def add_supplier(request):
    if request.method == 'POST':
        # Если форма была отправлена, обрабатываем данные
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()  # Сохраняем и получаем объект поставщика
            category_id = supplier.category.id  # Получаем ID города (категории)
            return redirect(f"{reverse('supplier_list')}?category={category_id}")  # Перенаправляем с фильтром
    else:
        form = SupplierForm()
    
    # Рендерим шаблон с формой
    return render(request, 'supplier/supplier_card.html', {'form': form})

@login_required
def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id) # Получаем поставщика по ID
    previous_category_id = supplier.category.id if supplier.category else None  # Сохраняем старый город

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier) # Заполняем форму данными поставщика
        if form.is_valid():
            supplier = form.save()  # Обновляем поставщика
            category_id = supplier.category.id if supplier.category else previous_category_id  # Берем новый город или старый
            return redirect(f"{reverse('supplier_list')}?category={category_id}") # Перенаправляем на список поставщиков
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'supplier/supplier_card.html', {'form': form})

@login_required
def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id) # Получаем поставщика по ID
    category_id = supplier.category.id if supplier.category else None  # Сохраняем город перед удалением

    if request.method == 'POST':
        supplier.delete() # Удаляем поставщика
        return redirect(f"{reverse('supplier_list')}?category={category_id}") # Перенаправляем на список поставщиков

    return render(request, 'supplier/confirm_delete_supplier.html', {'supplier': supplier})

@login_required
def get_supplier_token(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    token = SupplierToken.get_or_create_token(supplier)
    url = request.build_absolute_uri(f"/supplier_form/{supplier_id}/{token}/")
    return JsonResponse({"url": url})


def supplier_form(request, supplier_id, token):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    token_obj = get_object_or_404(SupplierToken, supplier=supplier, token=token)
    products = Product.objects.all()

    if token_obj.is_expired():
        return HttpResponseForbidden("Срок действия ссылки истек.")
    
    price_history = {}
    for product in products:
        history = Price.objects.filter(
            product=product,
            supplier=supplier
        ).order_by('-date_added')[:5]  # Последние 5 записей
        price_history[str(product.id)] = [
            {
                'price': item.price,
                'manufacturer': item.manufacturer,
                'date_added': item.date_added.strftime('%d.%m.%Y %H:%M')
            }
            for item in history
        ]

    if request.method == 'POST':
        print("POST Data:", request.POST)
        has_errors = False
        form_errors = {}  # Словарь для хранения ошибок по product_id
        form_data = {
            'prices': {},
            'manufacturers': {}
        }

        for product in products:
            product_id = str(product.id)
            price_str = request.POST.get(f'price_{product.id}', "").strip()
            manufacturer = request.POST.get(f'manufacturer_{product.id}', "").strip()
            
            # Сохраняем введенные данные
            form_data['prices'][product_id] = price_str
            form_data['manufacturers'][product_id] = manufacturer
            
            # Пропускаем если оба поля пустые
            if not price_str and not manufacturer:
                continue
                
            # Валидация цены
            price_error = None
            price = None
            
            if price_str:
                try:
                    price_str_clean = price_str.replace(" ", "").replace(",", ".")
                    if not all(c.isdigit() or c == '.' for c in price_str_clean):
                        raise ValueError("Некорректный формат числа")
                    
                    price = Decimal(price_str_clean)
                    
                    if price <= 0:
                        price_error = "Цена должна быть положительным числом"
                except (InvalidOperation, ValueError, TypeError) as e:
                    price_error = "Введите корректное число (разделитель дробной части - точка или запятая)"
            
            # Валидация производителя
            manufacturer_error = None
            if manufacturer and len(manufacturer) > 255:
                manufacturer_error = "Слишком длинное название (макс. 255 символов)"
            
            # Сохраняем ошибки
            if price_error or manufacturer_error:
                has_errors = True
                form_errors[product_id] = {
                    'price': price_error,
                    'manufacturer': manufacturer_error
                }
                continue
            
            # Всегда создаем новую запись, если есть данные (убрали проверку на существующую цену)
            try:
                Price.objects.create(
                    product=product,
                    supplier=supplier,
                    price=price,
                    manufacturer=manufacturer,
                    date_added=timezone.now()
                )
            except Exception as e:
                has_errors = True
                form_errors[product_id] = {
                    'general': f"Ошибка сохранения: {str(e)}"
                }

        if has_errors:
            return render(request, 'supplier_form/supplier_form.html', {
                "products": products,
                "supplier": supplier,
                "form_data": form_data,
                "form_errors": form_errors,
                "price_history": price_history  # Добавляем историю в контекст
            })
        
        messages.success(request, "Данные успешно сохранены!")
        return redirect('success')

    return render(request, 'supplier_form/supplier_form.html', {
        "products": products,
        "supplier": supplier,
        "form_data": None,
        "form_errors": None,
        "price_history": price_history  # Добавляем историю в контекст
    })

def success(request):
    return render(request, 'supplier_form/success.html')


def parse_time(time_str):
    """Парсит время из строки в формате HH:MM"""
    if not time_str:
        return None
    
    if re.match(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$', time_str):
        return datetime.strptime(time_str, '%H:%M').time()
    
    try:
        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1]) if len(parts) > 1 else 0
        return time(hours % 24, minutes % 60)
    except (ValueError, IndexError):
        return None

# @login_required
# def price_list(request):
#     date_str = request.GET.get('date')
#     time_from_str = request.GET.get('time_from')
#     time_to_str = request.GET.get('time_to')
#     supplier_id = request.GET.get('supplier')
    
#     prices = Price.objects.all()
    
#     if date_str:
#         try:
#             selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
#             time_from = parse_time(time_from_str)
#             time_to = parse_time(time_to_str)
            
#             if time_from:
#                 datetime_from = timezone.make_aware(datetime.combine(selected_date, time_from))
#                 prices = prices.filter(date_added__gte=datetime_from)
            
#             if time_to:
#                 datetime_to = timezone.make_aware(datetime.combine(selected_date, time_to))
#                 prices = prices.filter(date_added__lte=datetime_to)

#             if not time_from and not time_to:
#                 prices = prices.filter(date_added__date=selected_date)
#         except ValueError:
#             pass

#     if supplier_id:
#         prices = prices.filter(supplier_id=supplier_id)

#     suppliers = Supplier.objects.all()

#     return render(request, 'price_list.html', {
#         'prices': prices.order_by('-date_added'),
#         'selected_date': date_str,
#         'time_from': time_from_str,
#         'time_to': time_to_str,
#         'suppliers': suppliers,
#         'selected_supplier_id': supplier_id,
#     })

@login_required
def price_list(request):
    date_str = request.GET.get('date')
    time_from_str = request.GET.get('time_from')
    time_to_str = request.GET.get('time_to')
    supplier_id = request.GET.get('supplier')
    category_id = request.GET.get('category')  # Новое поле
    prices = Price.objects.select_related('product', 'supplier').all()

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            time_from = parse_time(time_from_str)
            time_to = parse_time(time_to_str)

            if time_from:
                datetime_from = timezone.make_aware(datetime.combine(selected_date, time_from))
                prices = prices.filter(date_added__gte=datetime_from)
            if time_to:
                datetime_to = timezone.make_aware(datetime.combine(selected_date, time_to))
                prices = prices.filter(date_added__lte=datetime_to)
            if not time_from and not time_to:
                prices = prices.filter(date_added__date=selected_date)
        except ValueError:
            pass

    if supplier_id:
        prices = prices.filter(supplier_id=supplier_id)

    if category_id:
        prices = prices.filter(supplier__category_id=category_id)

    suppliers = Supplier.objects.all()
    categories = Category.objects.all()  # Получаем все категории
    selected_category = category_id

    return render(request, 'price_list.html', {
        'prices': prices.order_by('-date_added'),
        'selected_date': date_str,
        'time_from': time_from_str,
        'time_to': time_to_str,
        'suppliers': suppliers,
        'selected_supplier_id': supplier_id,
        'categories': categories,
        'selected_category': selected_category,
    })

@login_required
def export_prices_to_excel(request):
    # Получаем параметры фильтрации
    date_str = request.GET.get('date')
    supplier_id = request.GET.get('supplier')

    # Получаем данные с фильтрацией
    prices = Price.objects.select_related('product', 'supplier').all()
    
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            prices = prices.filter(date_added__date=selected_date)
        except ValueError:
            pass
    
    if supplier_id:
        prices = prices.filter(supplier_id=supplier_id)

    # Создаем Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Цены"
    
    # Настройка страницы
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1

    # Стили для ячеек
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007bff", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки (соответствуют HTML-таблице)
    headers = [
        "Продукт",
        "Поставщик",
        "Цена",
        "Кол-во",
        "Ед. изм.",
        "Производитель",
        "Дата и время"
    ]
    sheet.append(headers)

    # Применяем стили к заголовкам
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Заполняем данными
    for price in prices:
        row = [
            str(price.product.name)[:20],  # Обрезаем до 20 символов как truncatechars
            str(price.supplier.name)[:20],
            float(price.price),  # Преобразуем Decimal к float для Excel
            int(price.product.quantity) if price.product.quantity else 0,
            str(price.product.unit),
            str(price.manufacturer)[:20] if price.manufacturer else "",
            timezone.localtime(price.date_added).strftime("%d.%m.%Y %H:%M:%S")  # Формат как в HTML
        ]
        sheet.append(row)

    # Применяем стили к данным
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Форматирование числовых полей
        row[2].number_format = '#,##0.00'  # Формат цены с 2 знаками после запятой
        row[3].number_format = '0'  # Целое число для количества

    # Автоподбор ширины колонок
    for col in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        column_letter = col[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)  # Ограничение максимальной ширины

    # Формируем имя файла
    filename = f"prices_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
    workbook.save(response)
    
    return response

# В views.py
@login_required
def product_supplier_table(request):
    products = Product.objects.all()
    suppliers = Supplier.objects.all()
    supplier_prices = Price.objects.select_related('supplier', 'product').all()

    # Фильтры
    date_str = request.GET.get('date')
    supplier_id = request.GET.get('supplier')
    category_id = request.GET.get('category')

    if date_str:
        selected_date = parse_date(date_str)
        supplier_prices = supplier_prices.filter(date_added__date=selected_date)

    if supplier_id:
        suppliers = suppliers.filter(id=supplier_id)
        supplier_prices = supplier_prices.filter(supplier_id=supplier_id)

    if category_id:
        suppliers = suppliers.filter(category_id=category_id)
        supplier_prices = supplier_prices.filter(supplier__category_id=category_id)

    # Исключаем поставщиков без данных
    active_suppliers = supplier_prices.values_list('supplier_id', flat=True).distinct()
    suppliers = suppliers.filter(id__in=active_suppliers)

    table_data = []
    for product in products:
        product_info = {
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'unit': product.unit,
            'supplier_prices': {},
            'best_price': None,
            'best_supplier': None,
            'categories': []  # Сюда добавим информацию по категориям
        }

        for supplier in suppliers:
            price_entry = supplier_prices.filter(
                product=product, 
                supplier=supplier
            ).order_by('-date_added').first()
            
            if price_entry:
                product_info['supplier_prices'][supplier.name] = {
                    'price': price_entry.price,
                    'manufacturer': price_entry.manufacturer
                }

        prices = {s: d['price'] for s, d in product_info['supplier_prices'].items() if d['price'] is not None}
        if prices:
            best_supplier_name = min(prices, key=prices.get)
            product_info['best_price'] = prices[best_supplier_name]
            product_info['best_supplier'] = best_supplier_name

        # Добавляем информацию по категориям и лучшим ценам (с учётом фильтров)
        categories = Category.objects.all()
        grouped = {}
        for category in categories:
            category_suppliers = suppliers.filter(category=category)
            category_prices = supplier_prices.filter(supplier__in=category_suppliers, product=product)

            if category_prices.exists():
                # Берём последнюю цену от каждого поставщика
                latest_prices = {}
                for sp in category_prices:
                    if sp.supplier.name not in latest_prices or sp.date_added > latest_prices[sp.supplier.name]['date']:
                        latest_prices[sp.supplier.name] = {
                            'price': sp.price,
                            'date': sp.date_added,
                            'supplier': sp.supplier,
                            'manufacturer': sp.manufacturer
                        }

                valid_prices = [item for item in latest_prices.values() if item['price'] is not None]

                if valid_prices:
                    best = min(valid_prices, key=lambda x: x['price'])
                    grouped[category.name] = {
                        'category': category.name,
                        'best_price': best['price'],
                        'best_supplier': best['supplier'].name,
                        'manufacturer': best['manufacturer']
                }
        product_info['categories'] = grouped
        table_data.append(product_info)

    categories = Category.objects.all()

    return render(request, 'product_supplier_table.html', {
        'table_data': table_data,
        'suppliers': suppliers,
        'categories': categories,
        'selected_date': date_str,
        'selected_supplier': supplier_id,
        'selected_category': category_id,
    })

@login_required
def best_prices_by_category(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    result = []

    for product in products:
        grouped = {}
        for category in categories:
            entries = Price.objects.filter(
                product=product, supplier__category=category
            ).select_related('supplier')

            if entries.exists():
                best = min(entries, key=lambda x: x.price)
                grouped[category.name] = {
                    'category': category,
                    'best_price': best.price,
                    'best_supplier': best.supplier,
                    'manufacturer': best.manufacturer
                }
        result.append({'product': product, 'categories': grouped})

    return render(request, 'best_prices_by_category.html', {'result': result})

@login_required
def price_history_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    prices = Price.objects.filter(product=product).select_related('supplier')

    # Получаем параметры фильтрации
    date_str = request.GET.get('date')
    supplier_id = request.GET.get('supplier')
    category_id = request.GET.get('category')

    if date_str:
        selected_date = parse_date(date_str)
        prices = prices.filter(date_added__date=selected_date)

    if supplier_id:
        prices = prices.filter(supplier_id=supplier_id)

    if category_id:
        prices = prices.filter(supplier__category_id=category_id)

    suppliers = Supplier.objects.all()
    categories = Category.objects.all()

    return render(request, 'price_history.html', {
        'product': product,
        'prices': prices.order_by('-date_added'),
        'suppliers': suppliers,
        'categories': categories,
        'selected_date': date_str,
        'selected_supplier': supplier_id,
        'selected_category': category_id,
    })
